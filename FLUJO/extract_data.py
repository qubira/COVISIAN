"""
Extrae datos de los Excel de DATA/COBERTURERO y DATA/PLANES hacia JSON
que consume el formulario en FLUJO/index.html.

Ejecutar de nuevo cada vez que se actualicen los Excel de origen:
    python extract_data.py
"""
import glob
import json
import os
import datetime
import openpyxl
import pyxlsb

ROOT = __file__.rsplit("\\", 2)[0] if "\\" in __file__ else __file__.rsplit("/", 2)[0]


def _ultimo_archivo(carpeta, patron):
    """Elige el archivo mas reciente (por fecha de modificacion) que matchee el patron
    dentro de la carpeta — asi cuando alguien pega un Excel nuevo (con nombre distinto,
    por ejemplo con la fecha en el nombre como los export de SAP) no hace falta tocar este
    script ni renombrar nada: el mas nuevo gana solo. Si hay varios, los viejos se pueden
    dejar ahi sin problema (no se leen), pero conviene borrarlos para no confundirse."""
    candidatos = [c for c in glob.glob(os.path.join(carpeta, patron))
                  if not os.path.basename(c).startswith("~$")]
    if not candidatos:
        raise FileNotFoundError("No se encontro ningun archivo '%s' en %s" % (patron, carpeta))
    return max(candidatos, key=os.path.getmtime)


COBERTURERO_PATH = _ultimo_archivo(ROOT + r"\DATA\COBERTURERO", "*.xlsx")
VISOR_PATH = _ultimo_archivo(ROOT + r"\DATA\PLANES", "*.xlsb")
STOCK_PATH = _ultimo_archivo(ROOT + r"\DATA\COLOR", "*.xlsx")
OUT_DIR = ROOT + r"\FLUJO\data"
EQUIPOS_IMG_DIR = ROOT + r"\DATA\EQUIPOS"
IMG_OUT_DIR = ROOT + r"\FLUJO\img\equipos"


def clean(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    if isinstance(v, str):
        return v.strip()
    return v


DIAS_ES = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"]


def _dia_entrega_real(fecha_venta, dias, valor_excel):
    """
    El Excel de origen (COBERTURERO) trae, en distintas filas y distintas versiones del
    archivo, DOS fórmulas distintas mezcladas para calcular la fecha/día de entrega:
    algunas filas usan WORKDAY.INTL(fecha, dias, 11) (salta domingo), otras una suma
    simple de fecha (+fecha+dias, sin saltar nada) que en la practica quedo mal copiada
    y da un dia antes cuando el rango cruza un domingo (confirmado comparando filas con
    la MISMA fecha de venta y mismos dias: las que usan WORKDAY.INTL y las que no
    difieren en exactamente 1 dia). Como el Excel se sigue actualizando y ese error de
    copiado puede reaparecer en cualquier fila/version futura, no confiamos en la columna
    ya calculada del Excel: la recalculamos aca mismo a partir de FECHA DE VENTA + DIAS
    (datos crudos, no formulas), aplicando la regla que SI se confirma correcta dentro
    del propio archivo (saltar solo domingo). Si fecha_venta/dias no son utilizables,
    se cae de vuelta al valor que ya trae el Excel en esa fila.

    OJO: fecha_venta es la celda =TODAY() cacheada por Excel la ultima vez que el archivo
    se guardo abierto en Excel (openpyxl con data_only=True lee ese valor congelado, no
    recalcula la formula) — NO es "hoy" en el momento en que un agente usa la app. Por eso
    el JSON tambien exporta "diasEntrega" (el numero crudo de dias) para que index.html
    recalcule el dia de entrega con la fecha real del navegador en cada carga, en vez de
    confiar en este valor ya congelado. Ver diaEntregaDesdeHoy() en index.html.
    """
    if not isinstance(fecha_venta, datetime.datetime) or not isinstance(dias, (int, float)):
        return clean(valor_excel)
    d = fecha_venta
    contador = 0
    dias_a_sumar = int(dias)
    while contador < dias_a_sumar:
        d = d + datetime.timedelta(days=1)
        if d.weekday() != 6:  # 6 = domingo
            contador += 1
    return DIAS_ES[d.weekday()]


def extract_cobertura():
    wb = openpyxl.load_workbook(COBERTURERO_PATH, read_only=True, data_only=True)
    rows = []

    ws = wb["LIMA Y CALLAO"]
    for r in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=18, values_only=True):
        if not r[0] or not r[2]:
            continue
        rows.append({
            "departamento": clean(r[0]),
            "provincia": clean(r[1]),
            "distrito": clean(r[2]),
            "express": clean(r[7]) or "NO",
            "hasta": clean(r[8]),
            "rango": clean(r[13]),
            "diaEntrega": _dia_entrega_real(r[3], r[14], r[17] or r[16]),
            "diasEntrega": int(r[14]) if isinstance(r[14], (int, float)) else None,
        })

    ws = wb["PROVINCIA"]
    for r in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=18, values_only=True):
        if not r[0] or not r[2]:
            continue
        rows.append({
            "departamento": clean(r[0]),
            "provincia": clean(r[1]),
            "distrito": clean(r[2]),
            "express": clean(r[7]) or "NO",
            "hasta": clean(r[8]),
            "rango": clean(r[11]),
            "diaEntrega": _dia_entrega_real(r[3], r[15], r[17]),
            "diasEntrega": int(r[15]) if isinstance(r[15], (int, float)) else None,
        })

    # dedupe by distrito+provincia+departamento (keep first)
    seen = set()
    uniq = []
    for row in rows:
        key = (row["departamento"], row["provincia"], row["distrito"])
        if key in seen:
            continue
        seen.add(key)
        uniq.append(row)

    uniq.sort(key=lambda x: (x["departamento"] or "", x["provincia"] or "", x["distrito"] or ""))
    return uniq


def extract_specs():
    wb = openpyxl.load_workbook(COBERTURERO_PATH, read_only=True, data_only=True)
    ws = wb["Base"]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=14, values_only=True))
    out = []
    for r in rows:
        if not r[1]:
            continue
        out.append({
            "gama": clean(r[0]),
            "equipo": clean(r[1]),
            "descAgrup": clean(r[2]),
            "pEspecial": clean(r[3]),
            "marca": clean(r[4]),
            "pantalla": clean(r[5]),
            "camaraPrincipal": clean(r[6]),
            "camaraSelfie": clean(r[7]),
            "memoria": clean(r[8]),
            "ram": clean(r[9]),
            "bateria": clean(r[10]),
            "so": clean(r[11]),
            "procesador": clean(r[12]),
            "precioAprox": clean(r[13]),
        })
    return out


def extract_equipos_planes():
    names = set()
    with pyxlsb.open_workbook(VISOR_PATH) as wb:
        with wb.get_sheet("BaseTerminales") as sheet:
            rows = list(sheet.rows())
        for r in rows[1:]:
            if r and r[0].v:
                name = str(r[0].v).strip()
                if name:
                    names.add(name)
    return sorted(names)


def extract_gama_financiamiento():
    """
    Tramo de gama (1-7, columna "grupo_fin" de BaseTerminales) por equipo, solo de las
    filas de financiamiento a "12 cuotas" (es el tramo que usa el simulador VBA de
    Visor.xlsb para fijar la cuota inicial de los CÓD 1 y 5 de financiamiento: cada tramo
    corresponde a un monto de inicial fijo, no a un porcentaje).
    Se probo que un mismo equipo no trae tramos distintos entre CAEQ/MIGRA y PORTA (0
    conflictos sobre 198 equipos), asi que se guarda como mapeo plano equipo -> tramo.
    """
    datos = {}
    with pyxlsb.open_workbook(VISOR_PATH) as wb:
        with wb.get_sheet("BaseTerminales") as sheet:
            rows = list(sheet.rows())
    for r in rows[1:]:
        if not r or not r[0].v:
            continue
        tipo_trans = r[3].v if len(r) > 3 else None
        if tipo_trans != "12 cuotas":
            continue
        tramo = r[9].v if len(r) > 9 else None
        if tramo is None:
            continue
        nombre = str(r[0].v).strip().upper()
        if not nombre:
            continue
        datos[nombre] = int(tramo)
    return datos


def extract_equipos_tipoventa():
    """
    Para cada terminal de BaseTerminales, junta los valores de TipoVenta (columna 2:
    CAEQ/MIGRA, PORTA) y canal (columna 8: ESP_POST_PHOENIX, ESP_POST_ONLINE,
    ESP_POST_RETENCION, ESP_POST_BLINDAJE) que aparecen en sus distintas filas.
    Un mismo terminal aparece varias veces (una fila por combinacion de TipoVenta/canal
    en la que se vende), asi que se acumulan en un set por terminal.

    OJO: esta info es PARCIAL. No hay ningun tag para "ALTA" ni para "PREPAGO" en el
    archivo fuente, asi que el filtro de TIPO/SUBTIPO en el formulario solo aplica para
    CAEQ, PORTABILIDAD y ESPECIALES (Phoenix/Online/Retencion-Blindaje); para ALTA,
    PREPAGO y OTROS no hay como filtrar sin arriesgar ocultar equipos validos.
    """
    acumulado = {}
    with pyxlsb.open_workbook(VISOR_PATH) as wb:
        with wb.get_sheet("BaseTerminales") as sheet:
            rows = list(sheet.rows())
    for r in rows[1:]:
        if not r or not r[0].v:
            continue
        nombre = str(r[0].v).strip()
        if not nombre:
            continue
        tipo_venta = r[2].v if len(r) > 2 else None
        canal = r[8].v if len(r) > 8 else None
        entry = acumulado.setdefault(nombre, {"tipoVenta": set(), "canal": set()})
        if tipo_venta:
            entry["tipoVenta"].add(str(tipo_venta).strip())
        if canal:
            entry["canal"].add(str(canal).strip())
    return {
        k: {"tipoVenta": sorted(v["tipoVenta"]), "canal": sorted(v["canal"])}
        for k, v in acumulado.items()
        if v["tipoVenta"] or v["canal"]
    }


def extract_stock():
    """
    Stock por ubicacion (INTSAP52). Las hojas del libro representan zonas:
    LIMA -> departamento LIMA (incluye CALLAO como provincia)
    AREQUIPA / LA LIBERTAD / LAMBAYEQUE / PIURA -> ese departamento puntual
    PROVINCIAS -> el resto de departamentos del pais
    """
    wb = openpyxl.load_workbook(STOCK_PATH, read_only=True, data_only=True)
    out = []
    for zona in wb.sheetnames:
        ws = wb[zona]
        for r in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=3, values_only=True):
            if not r[1]:
                continue
            out.append({
                "zona": zona,
                "codigoSap": clean(r[0]),
                "modelo": clean(r[1]),
                "estado": clean(r[2]),
            })
    return out


def extract_packs():
    """
    Combina las 3 hojas de packs/combos de COBERTURERO en un solo listado,
    cada una etiquetada por su origen/version.
    """
    wb = openpyxl.load_workbook(COBERTURERO_PATH, read_only=True, data_only=True)
    fuentes = [
        ("PACKS", "PACKS (version marzo 2026)"),
        ("PACK  ", "PACK (version mayo 2026)"),
        ("PACK", "CAMPANA (version mayo 2026)"),
    ]
    out = []
    for sheet_name, etiqueta in fuentes:
        ws = wb[sheet_name]
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6, values_only=True):
            marca, equipo, nombreStock, accesorio, precio = r[1], r[2], r[3], r[4], r[5]
            if marca in (None, "MARCA") and equipo in (None, "EQUIPO"):
                continue
            if not equipo and not nombreStock:
                continue
            if not accesorio:
                continue
            out.append({
                "origen": etiqueta,
                "marca": clean(marca),
                "equipo": clean(equipo),
                "nombreStock": clean(nombreStock),
                "accesorio": clean(accesorio),
                "precio": clean(precio),
            })
    return out


def extract_planes():
    """
    Precios de plan "Mi Movistar" (columnas de rango de precio) usados en las hojas de
    Visor.xlsb (ALTA/PORTA/CAEQ/ESP_*). CAEQ_GENERICA trae la lista mas completa.
    """
    planes = []
    with pyxlsb.open_workbook(VISOR_PATH) as wb:
        with wb.get_sheet("CAEQ_GENERICA") as sheet:
            rows = list(sheet.rows())
    header = [c.v for c in rows[6]]  # fila 7 (0-indexed 6): Categoria, Modelo, precios...
    vistos = set()
    for v in header:
        if isinstance(v, (int, float)) and v not in vistos:
            vistos.add(v)
            planes.append(round(v, 1))
        elif isinstance(v, str) and v.strip().upper() == "MT" and "MT" not in vistos:
            vistos.add("MT")
            planes.append("MT")
    return planes


def _leer_bloques_hoja_precios(wb, sheet_name):
    """
    Algoritmo generico para cualquier hoja de precios de Visor.xlsb:
    1. Ubica la celda con valor "Modelo" (busca en todas las filas/columnas) -> esa fila
       es la fila de encabezado de precios, esa columna es la columna de nombre de equipo.
    2. Las columnas a la derecha de "Modelo" en esa misma fila son los "planes" (precio
       Mi Movistar, numerico, o "MT"/"Planes MT" para linea nueva).
    3. Como una misma hoja puede traer varios bloques de columnas en paralelo (ej. CONTADO
       y FINANCIADO uno al lado del otro), se detecta un bloque nuevo cada vez que se repite
       una etiqueta de plan ya vista (cada bloque recorre el mismo set ordenado de planes).
    Devuelve una lista de bloques: [{"col_inicio": int, "planes": [...], "precios": {modelo: {plan: costo}}}]
    """
    with wb.get_sheet(sheet_name) as sheet:
        rows = list(sheet.rows())

    header_row_idx = None
    modelo_col = None
    for i, row in enumerate(rows):
        for c in row:
            if c.v and isinstance(c.v, str) and c.v.strip().lower() == "modelo":
                header_row_idx = i
                modelo_col = c.c
                break
        if header_row_idx is not None:
            break
    if header_row_idx is None:
        return []

    header_map = {c.c: c.v for c in rows[header_row_idx]}
    max_col = max(header_map.keys())

    price_cols = []
    for col in range(modelo_col + 1, max_col + 1):
        v = header_map.get(col)
        if v is None:
            continue
        if isinstance(v, (int, float)):
            price_cols.append((col, round(float(v), 1)))
        elif isinstance(v, str) and "mt" in v.strip().lower():
            price_cols.append((col, "MT"))

    bloques_cols = []
    bloque_actual = []
    vistos = set()
    for col, label in price_cols:
        if label in vistos:
            bloques_cols.append(bloque_actual)
            bloque_actual = []
            vistos = set()
        vistos.add(label)
        bloque_actual.append((col, label))
    if bloque_actual:
        bloques_cols.append(bloque_actual)

    filas_data = rows[header_row_idx + 1:]
    bloques = []
    for cols in bloques_cols:
        precios = {}
        for row in filas_data:
            row_map = {c.c: c.v for c in row}
            modelo = row_map.get(modelo_col)
            if not modelo or not isinstance(modelo, str) or not modelo.strip():
                continue
            modelo = clean(modelo).upper()
            for col, label in cols:
                val = row_map.get(col)
                if isinstance(val, (int, float)) and val > 0:
                    precios.setdefault(modelo, {})[str(label)] = round(float(val), 2)
        bloques.append({
            "col_inicio": cols[0][0],
            "planes": [l for _, l in cols],
            "precios": precios,
        })
    return bloques


# Mapea cada combinacion TIPO|SUBTIPO del formulario a la(s) hoja(s) de Visor.xlsb que
# traen su precio, y a que "convenio" (posicion de bloque dentro de la hoja) corresponde
# cada uno. bloque_idx es el indice del bloque tal como aparece de izquierda a derecha en
# la hoja (0 = primer bloque de columnas, 1 = segundo, etc), confirmado leyendo las
# etiquetas CONTADO/FINANCIADO/etc que Visor.xlsb trae en la fila justo arriba de "Modelo".
_MAPA_PRECIOS_TIPO_SUBTIPO = {
    "POSTPAGO|ALTA": [
        {"hoja": "ALTA PP y PF", "bloque_idx": 0, "convenio": "CONTADO"},
    ],
    "POSTPAGO|PORTABILIDAD": [
        {"hoja": "PORTA_GENERICA", "bloque_idx": 0, "convenio": "CONTADO"},
        {"hoja": "PORTA_GENERICA", "bloque_idx": 1, "convenio": "FINANCIADO"},
    ],
    "POSTPAGO|CAEQ": [
        {"hoja": "CAEQ_GENERICA", "bloque_idx": 0, "convenio": "CONTADO"},
        {"hoja": "CAEQ_GENERICA", "bloque_idx": 1, "convenio": "FINANCIADO"},
    ],
    "ESPECIALES|PHOENIX (Contado)": [
        {"hoja": "ESP_PHOENIX", "bloque_idx": 0, "convenio": "CONTADO"},
    ],
    "ESPECIALES|ONLINE (Contado)": [
        {"hoja": "ESP_ONLINE", "bloque_idx": 0, "convenio": "CONTADO"},
    ],
    # ESP_BLINDAJE (Contado) es el par exacto de ESP_BLINDAJE_FIN (mismo titulo "RETENCION/
    # BLINDAJE", el otro solo Contado). ESP_RETENCION trae otro listado de equipos para el
    # mismo caso (titulo identico "CONTADO: RETENCION/BLINDAJE"): se agrega como bloque
    # adicional del mismo convenio CONTADO para no perder equipos que solo estan en una hoja
    # (si un equipo aparece en las dos, gana la primera = ESP_BLINDAJE (Contado)).
    "ESPECIALES|RETENCION/BLINDAJE (Contado)": [
        {"hoja": "ESP_BLINDAJE (Contado)", "bloque_idx": 0, "convenio": "CONTADO"},
        {"hoja": "ESP_RETENCION", "bloque_idx": 0, "convenio": "CONTADO"},
    ],
    "ESPECIALES|PHOENIX/ONLINE (Financiado)": [
        {"hoja": "ESP_PHOENIX_FIN", "bloque_idx": 0, "convenio": "FINANCIADO"},
    ],
    "ESPECIALES|RETENCION/BLINDAJE (Financiado)": [
        {"hoja": "ESP_BLINDAJE_FIN", "bloque_idx": 0, "convenio": "FINANCIADO"},
    ],
    # BANCOS (cuotas sin intereses) es transversal a ALTA/PORTA/CAEQ y no usa CONTADO/
    # FINANCIADO sino dos vistas del mismo precio: total y cuota mensual.
    "OTROS|BANCOS (Cuotas sin intereses)": [
        {"hoja": "MP-Cuotas sin intereses", "bloque_idx": 0, "convenio": "PRECIO_TOTAL"},
        {"hoja": "MP-Cuotas sin intereses", "bloque_idx": 1, "convenio": "CUOTA_MENSUAL"},
    ],
    # POS1 (Clientes Black) solo aplica para CAEQ; trae 3 bloques: contado, financiado
    # propio de la tienda, y financiado via bancos.
    "OTROS|CLIENTES BLACK (Dscto Equipos)": [
        {"hoja": "POS1", "bloque_idx": 0, "convenio": "CONTADO"},
        {"hoja": "POS1", "bloque_idx": 1, "convenio": "FINANCIADO_PROPIO"},
        {"hoja": "POS1", "bloque_idx": 2, "convenio": "BANCOS"},
    ],
    # PREPAGO no tiene mapeo: CAM_PREPAGO esta indexada por tipo de transaccion
    # (ALTA/PORTA/CAEQ), no por SUBTIPO=PREPAGO, asi que no hay forma confiable de
    # cruzarla sin arriesgar mostrar un precio equivocado. El formulario debe avisar
    # "no disponible para PREPAGO" en vez de adivinar.
}


def extract_precios():
    """
    Precios de equipos por TIPO|SUBTIPO + convenio (CONTADO/FINANCIADO/etc), usados por
    la seccion COSTO del formulario. Estructura del JSON resultante:
    {
      "POSTPAGO|CAEQ": {
        "convenios": ["CONTADO", "FINANCIADO"],
        "planes": {"CONTADO": [29.9, 35.9, ..., "MT"], "FINANCIADO": [...]},
        "precios": {"CONTADO": {"IPHONE 13 128GB": {"29.9": 1989.0, ...}}, "FINANCIADO": {...}}
      },
      ...
    }
    Cuando un TIPO|SUBTIPO tiene mas de un bloque fuente para el mismo convenio (caso
    RETENCION/BLINDAJE Contado), se fusionan: el primer bloque manda si un equipo aparece
    en ambos.
    """
    resultado = {}
    with pyxlsb.open_workbook(VISOR_PATH) as wb:
        cache_hojas = {}
        for clave, fuentes in _MAPA_PRECIOS_TIPO_SUBTIPO.items():
            convenios_data = {}
            for fuente in fuentes:
                hoja = fuente["hoja"]
                if hoja not in cache_hojas:
                    cache_hojas[hoja] = _leer_bloques_hoja_precios(wb, hoja)
                bloques = cache_hojas[hoja]
                if fuente["bloque_idx"] >= len(bloques):
                    continue
                bloque = bloques[fuente["bloque_idx"]]
                convenio = fuente["convenio"]
                destino = convenios_data.setdefault(convenio, {"planes": bloque["planes"], "precios": {}})
                for modelo, precios_modelo in bloque["precios"].items():
                    destino["precios"].setdefault(modelo, precios_modelo)
            if not convenios_data:
                continue
            resultado[clave] = {
                "convenios": list(convenios_data.keys()),
                "planes": {c: d["planes"] for c, d in convenios_data.items()},
                "precios": {c: d["precios"] for c, d in convenios_data.items()},
            }
    return resultado


def extract_cod_financiamiento():
    codes = set()
    sheets = ["ALTA PP y PF", "PORTA_GENERICA", "CAEQ_GENERICA", "MP-Cuotas sin intereses",
              "ESP_PHOENIX", "ESP_PHOENIX_FIN", "ESP_BLINDAJE (Contado)", "ESP_BLINDAJE_FIN",
              "ESP_ONLINE", "ESP_RETENCION", "POS1"]
    with pyxlsb.open_workbook(VISOR_PATH) as wb:
        for name in sheets:
            with wb.get_sheet(name) as sheet:
                rows = list(sheet.rows())
            if rows:
                for c in rows[0]:
                    if isinstance(c.v, (int, float)):
                        codes.add(int(c.v))
    return sorted(codes)


# Excel de origen cuyos cambios deben disparar una re-extraccion. Si se agrega un excel
# nuevo como fuente de datos, agregarlo aca tambien para que quede cubierto por la deteccion
# automatica de "datos desactualizados" (INICIAR.bat / _necesita_actualizar).
ARCHIVOS_ORIGEN = {
    "cobertero": COBERTURERO_PATH,
    "visor": VISOR_PATH,
    "stock": STOCK_PATH,
}
META_PATH_REL = r"\_meta.json"


def _escribir_meta():
    import os
    meta = {
        "generadoEn": datetime.datetime.now().isoformat(timespec="seconds"),
        "fuentes": {
            nombre: os.path.getmtime(ruta)
            for nombre, ruta in ARCHIVOS_ORIGEN.items()
            if os.path.exists(ruta)
        },
    }
    with open(OUT_DIR + META_PATH_REL, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=0)


def necesita_actualizar():
    """
    True si algun Excel de origen se modifico despues de la ultima extraccion (o si
    todavia no se extrajo nunca). Se usa desde INICIAR.bat para regenerar los JSON
    solo cuando realmente hace falta, sin que el agente tenga que acordarse de hacerlo
    a mano cada vez que edita un Excel.
    """
    import os
    meta_path = OUT_DIR + META_PATH_REL
    if not os.path.exists(meta_path):
        return True
    try:
        with open(meta_path, encoding="utf-8") as f:
            meta = json.load(f)
    except (OSError, ValueError):
        return True
    fuentes_guardadas = meta.get("fuentes", {})
    for nombre, ruta in ARCHIVOS_ORIGEN.items():
        if not os.path.exists(ruta):
            continue
        mtime_actual = os.path.getmtime(ruta)
        mtime_guardado = fuentes_guardadas.get(nombre)
        if mtime_guardado is None or mtime_actual > mtime_guardado:
            return True
    return False


def _slug_imagen(nombre_sin_ext):
    """
    "MOTO EDGE 70 FUSION correct.png" / "Samsung S26 Ultra_.jpg" -> nombre de archivo
    limpio para URL (sin espacios ni mayusculas mezcladas) + nombre legible para el
    matching de tokens en el frontend (mismo criterio que ya limpia nombres de equipo
    en otras extracciones de este archivo: quita basura de nomenclatura interna, no
    cambia el equipo en si).
    """
    import re
    s = re.sub(r"\bcorrect\b", "", nombre_sin_ext, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip().strip("_").strip()
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return slug, s


def extract_equipos_imagenes():
    """
    Copia las fotos de DATA/EQUIPOS (una foto por linea de modelo, no por variante de
    almacenamiento/color) a FLUJO/img/equipos con nombres de archivo limpios, y arma el
    manifest equipos_imagenes.json que el frontend usa para encontrarle la foto a cada
    equipo del catalogo por matching de tokens (misma logica que ya usa mejorMatchSpecs
    en index.html para la ficha tecnica).
    """
    import os
    import shutil

    if not os.path.isdir(EQUIPOS_IMG_DIR):
        return []

    os.makedirs(IMG_OUT_DIR, exist_ok=True)
    # limpia el destino primero para no dejar fotos huerfanas si se borra/renombra el
    # archivo de origen entre una extraccion y la siguiente
    for f in os.listdir(IMG_OUT_DIR):
        os.remove(os.path.join(IMG_OUT_DIR, f))

    manifest = []
    usados = set()
    for nombre_archivo in sorted(os.listdir(EQUIPOS_IMG_DIR)):
        ruta = os.path.join(EQUIPOS_IMG_DIR, nombre_archivo)
        if not os.path.isfile(ruta):
            continue
        base, ext = os.path.splitext(nombre_archivo)
        ext = ext.lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp"):
            continue
        slug, nombre_limpio = _slug_imagen(base)
        if not slug:
            continue
        archivo_final = slug
        n = 2
        while archivo_final in usados:
            archivo_final = slug + "-" + str(n)
            n += 1
        usados.add(archivo_final)
        shutil.copyfile(ruta, os.path.join(IMG_OUT_DIR, archivo_final + ext))
        manifest.append({"nombre": nombre_limpio.upper(), "archivo": "img/equipos/" + archivo_final + ext})

    return manifest


def generar_cambios_historicos(stock_nuevo):
    """
    Compara el stock recien extraido contra el stock.json que habia ANTES de esta corrida
    (osea la extraccion anterior), para el boton "Cambios historicos" del formulario. Hay que
    llamarla con el stock.json viejo todavia en disco, antes de que main() lo sobreescriba.

    La clave de comparacion es (zona, modelo) y NO solo modelo: el mismo equipo+color puede
    existir en varias zonas (LIMA, AREQUIPA, PROVINCIAS, etc. — son las pestañas del Excel de
    DATA/COLOR) con stock independiente en cada una. Comparar solo por modelo escondia el caso
    real de "se fue de LIMA pero sigue en AREQUIPA" (no se veia como cambio porque el modelo
    seguia existiendo en ALGUNA zona) — con (zona, modelo) cada zona se seria un cambio aparte.

    Si nunca hubo una extraccion anterior (primera vez que corre este script en la maquina)
    no hay con que comparar, asi que no se reporta ningun cambio (evita el falso positivo de
    marcar TODO el stock como "agregado").
    """
    ruta_stock_viejo = OUT_DIR + r"\stock.json"
    ruta_meta_vieja = OUT_DIR + META_PATH_REL
    ruta_cambios = OUT_DIR + r"\cambios_historicos.json"

    if not os.path.exists(ruta_stock_viejo):
        return {"fechaAnterior": None, "fechaActual": datetime.datetime.now().isoformat(timespec="seconds"),
                "agregados": [], "eliminados": []}

    with open(ruta_stock_viejo, encoding="utf-8") as f:
        claves_viejas = set((row["zona"], row["modelo"]) for row in json.load(f) if row.get("modelo"))

    fecha_anterior = None
    if os.path.exists(ruta_meta_vieja):
        try:
            with open(ruta_meta_vieja, encoding="utf-8") as f:
                fecha_anterior = json.load(f).get("generadoEn")
        except Exception:
            pass

    claves_nuevas = set((row["zona"], row["modelo"]) for row in stock_nuevo if row.get("modelo"))
    agregados = [{"zona": z, "modelo": m} for z, m in sorted(claves_nuevas - claves_viejas)]
    eliminados = [{"zona": z, "modelo": m} for z, m in sorted(claves_viejas - claves_nuevas)]

    # Si esta corrida no trajo ningun cambio real de stock (el Excel de COLOR no cambio,
    # solo se re-corrio el script por otra fuente), se conserva el ultimo diff con cambios
    # reales que haya en disco, en vez de pisarlo con un diff vacio — asi el boton siempre
    # muestra la ultima actualizacion real de stock, no "sin cambios" por casualidad de cuando
    # se corrio el script.
    if not agregados and not eliminados and os.path.exists(ruta_cambios):
        try:
            with open(ruta_cambios, encoding="utf-8") as f:
                anterior = json.load(f)
            if anterior.get("agregados") or anterior.get("eliminados"):
                return anterior
        except Exception:
            pass

    return {
        "fechaAnterior": fecha_anterior,
        "fechaActual": datetime.datetime.now().isoformat(timespec="seconds"),
        "agregados": agregados,
        "eliminados": eliminados,
    }


def main():
    import os
    os.makedirs(OUT_DIR, exist_ok=True)

    cobertura = extract_cobertura()
    specs = extract_specs()
    equipos_planes = extract_equipos_planes()
    cod_fin = extract_cod_financiamiento()
    gama_fin = extract_gama_financiamiento()
    stock = extract_stock()
    cambios_historicos = generar_cambios_historicos(stock)
    packs = extract_packs()
    planes = extract_planes()
    equipos_tipoventa = extract_equipos_tipoventa()
    precios = extract_precios()
    equipos_imagenes = extract_equipos_imagenes()

    with open(OUT_DIR + r"\cobertura.json", "w", encoding="utf-8") as f:
        json.dump(cobertura, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\equipos_specs.json", "w", encoding="utf-8") as f:
        json.dump(specs, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\equipos_planes.json", "w", encoding="utf-8") as f:
        json.dump(equipos_planes, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\cod_financiamiento.json", "w", encoding="utf-8") as f:
        json.dump(cod_fin, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\gama_financiamiento.json", "w", encoding="utf-8") as f:
        json.dump(gama_fin, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\stock.json", "w", encoding="utf-8") as f:
        json.dump(stock, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\packs.json", "w", encoding="utf-8") as f:
        json.dump(packs, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\planes.json", "w", encoding="utf-8") as f:
        json.dump(planes, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\equipos_tipoventa.json", "w", encoding="utf-8") as f:
        json.dump(equipos_tipoventa, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\precios.json", "w", encoding="utf-8") as f:
        json.dump(precios, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\equipos_imagenes.json", "w", encoding="utf-8") as f:
        json.dump(equipos_imagenes, f, ensure_ascii=False, indent=0)
    with open(OUT_DIR + r"\cambios_historicos.json", "w", encoding="utf-8") as f:
        json.dump(cambios_historicos, f, ensure_ascii=False, indent=0)

    _escribir_meta()

    print("planes:", planes)
    print("equipos_tipoventa:", len(equipos_tipoventa), "equipos con tag de TipoVenta/canal")
    print("cobertura:", len(cobertura), "distritos")
    print("specs:", len(specs), "equipos con ficha tecnica")
    print("equipos_planes:", len(equipos_planes), "equipos en visor de precios")
    print("cod_financiamiento:", cod_fin)
    print("gama_financiamiento:", len(gama_fin), "equipos con tramo de gama para inicial fija")
    print("stock:", len(stock), "filas de stock por zona")
    print("packs:", len(packs), "combos con accesorio")
    print("precios:", len(precios), "combinaciones TIPO|SUBTIPO con tabla de precios")
    print("equipos_imagenes:", len(equipos_imagenes), "fotos de equipos copiadas a FLUJO/img/equipos")
    print("cambios_historicos:", len(cambios_historicos["agregados"]), "agregados,",
          len(cambios_historicos["eliminados"]), "eliminados desde", cambios_historicos["fechaAnterior"])


if __name__ == "__main__":
    import sys
    if "--si-hace-falta" in sys.argv:
        # Usado por INICIAR.bat en cada arranque: si ningun Excel de origen cambio desde
        # la ultima extraccion, no vuelve a leer los Excel (rapido); si alguno cambio
        # (o nunca se extrajo), corre la extraccion completa como siempre.
        if necesita_actualizar():
            print("Datos desactualizados o inexistentes: extrayendo de los Excel...")
            main()
        else:
            print("Los datos ya estan al dia con los Excel de origen, no hace falta releerlos.")
    else:
        main()
